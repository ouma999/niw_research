"""
Builds environment/data/assumptions.xlsx -- the second messy source
artifact. The Mortality_Table tab is a decoy: it shows a table version
code change (which reads as "something changed here") but the change
has already been fully reflected in the given interest_cost and
expected_return figures in instruction.md, so no further adjustment is
needed for it. A model that reflexively tries to "adjust for" the
mortality code change is applying a plausible-looking but unnecessary
correction.
"""
from pathlib import Path
from openpyxl import Workbook

XLSX_PATH = Path(__file__).resolve().parent.parent / "environment" / "data" / "assumptions.xlsx"


def build():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Discount_Rate"
    ws1.append(["Period", "Discount Rate"])
    ws1.append(["Prior period", 0.0525])
    ws1.append(["Current period", 0.0460])

    ws2 = wb.create_sheet("Mortality_Table")
    ws2.append(["Period", "Table Version Code"])
    ws2.append(["Prior period", "RP-2014-TotalDataset-MP2019"])
    ws2.append(["Current period", "RP-2014-TotalDataset-MP2021"])

    ws3 = wb.create_sheet("Expected_Return")
    ws3.append(["Period", "Expected Long-Term Return on Plan Assets"])
    ws3.append(["Prior period", 0.0550])
    ws3.append(["Current period", 0.0550])

    ws4 = wb.create_sheet("Amortization_Schedule")
    ws4.append(["Component", "Current Period Amount"])
    ws4.append(["Amortization of net actuarial loss", 52000.00])

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


if __name__ == "__main__":
    build()
